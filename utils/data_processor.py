"""
Core data processing utilities for Excel file manipulation.

This module handles reading, processing, and writing Excel files for
the ROESAN-SALVAGUARDAR reconciliation application.
"""

import os
import shutil
import pandas as pd
import numpy as np
from datetime import datetime
import logging
from typing import Dict, List, Any, Tuple, Optional, Set
from openpyxl import load_workbook

import config
from .validators import clean_id, split_name

logger = logging.getLogger(__name__)


def process_excel_files(
    master_path: str,
    movement_paths: List[str],
    cobro_mes: int = 0,
    cobro_anio: int = 0,
    header_search_rows: int = 20
) -> Dict[str, Any]:
    """
    Process master template and movement files for reconciliation.

    Reads the master template, extracts insured individuals, then processes
    movement files (INGRESOS/RETIROS) to generate reconciliation results.

    Args:
        master_path: Path to master template file (PLANTILLA CARGUE COBRO)
        movement_paths: List of movement file paths containing INGRESOS/RETIROS
        header_search_rows: Number of rows to search for headers (default: 20)

    Returns:
        Dict[str, Any]: Processing results containing:
            - status: 'success' or 'error'
            - statistics: Movement counts and master row counts
            - dataframes: Processed movements and master DataFrames
            - preview: Preview data for UI
            - logs: Processing log messages
            - error: Error message if status is 'error'

    Raises:
        Exception: If master file cannot be read or processing fails
    """
    logs: List[str] = []
    processed_rows: List[Dict[str, Any]] = []
    all_ingresos_ids: Set[str] = set()
    all_retiros_ids: Set[str] = set()

    # 1. Read Master Template
    try:
        master_df: pd.DataFrame = _read_excel_safe(master_path)
    except FileNotFoundError as e:
        raise FileNotFoundError(f'Master file not found: {master_path}')
    except pd.errors.ParserError as e:
        raise ValueError(f'Invalid Excel format in master file: {e}')
    except Exception as e:
        raise Exception(f'Error reading master file: {e}')

    # Build ID lookup from master
    id_lookup: Dict[str, Dict[str, Any]] = {}
    for idx, row in master_df.iterrows():
        cid: Optional[str] = clean_id(row.get('NUMERO_IDENTIFICACION_ASEGURADO'))
        if cid:
            id_lookup[cid] = {
                'CARGO': row.get('CARGO'),
                'GENERO': _normalize_genero(row.get('GENERO')),
                'FECHA_NACIMIENTO_ASEGURADO': row.get('FECHA_NACIMIENTO_ASEGURADO'),
                'VALOR_ASEGURADO_MUERTE': row.get('VALOR_ASEGURADO_MUERTE', 5000000)
            }

    logs.append(f"Master file loaded: {len(id_lookup)} unique IDs")

    # 2. Process Movement Files
    logs.append(f"Processing {len(movement_paths)} movement files...")

    for fpath in movement_paths:
        fname = os.path.basename(fpath)

        try:
            xl = pd.ExcelFile(fpath)
        except Exception as e:
            logs.append(f"[!] Error reading {fname}: {e}")
            continue

        for sheet_name in ['INGRESOS', 'RETIROS']:
            if sheet_name not in xl.sheet_names:
                continue

            logs.append(f"Processing {fname} - Sheet: {sheet_name}")

            try:
                df_raw = pd.read_excel(fpath, sheet_name=sheet_name, header=None)

                # Locate header row
                header_idx = _find_header_row(df_raw, header_search_rows)

                if header_idx == -1:
                    logs.append(f"  [!] Warning: Headers not found in {fname} - {sheet_name}")
                    continue

                # Extract data
                df = df_raw.iloc[header_idx+1:].copy()
                df.columns = df_raw.iloc[header_idx].values

                # Standardize column names
                cols = [str(c).upper().strip() for c in df.columns]
                df.columns = cols

                # Find relevant columns
                id_col = _find_column(cols, ['IDENTIFICACION', 'C.C.', 'CEDULA', 'DOC'])
                name_col = _find_column(cols, ['NOMBRE', 'RAZON SOCIAL', 'APELLIDO'])
                fecha_nac_col = _find_column(cols, ['NACIMIENTO'])
                # Look for movement date column - prioritize INGRESO/RETIRO which appear in "FECHA DE INGRESO/RETIRO"
                fecha_nov_col = None
                # First try to find "FECHA DE INGRESO" or "FECHA DE RETIRO"
                for col in cols:
                    col_upper = col.upper() if col else ''
                    if 'INGRESO' in col_upper and 'FECHA' in col_upper:
                        fecha_nov_col = col
                        break
                    if 'RETIRO' in col_upper and 'FECHA' in col_upper:
                        fecha_nov_col = col
                        break
                # Fall back to looking for other keywords
                if not fecha_nov_col:
                    fecha_nov_col = _find_column(cols, ['NOVEDAD', 'INGRESO', 'RETIRO', 'EGRESO', 'MOVIMIENTO'])
                # Last resort: any FECHA column
                if not fecha_nov_col:
                    fecha_nov_col = _find_column(cols, ['FECHA'])

                if not id_col or not name_col:
                    logs.append(f"  [!] Error: Required columns not found in {fname} - {sheet_name}")
                    continue

                # Process rows
                count_records = 0
                for _, row in df.iterrows():
                    n_id = clean_id(row.get(id_col))
                    if not n_id or n_id == 'NAN':
                        continue

                    full_name = row.get(name_col)
                    is_ingreso = sheet_name == 'INGRESOS'
                    apellidos, nombres = split_name(full_name)

                    # Get info from master lookup
                    lookup_data = id_lookup.get(n_id, {})
                    genero = _normalize_genero(lookup_data.get('GENERO'))
                    valor_aseg = lookup_data.get('VALOR_ASEGURADO_MUERTE', 5000000)
                    # CARGO: always "AGENTES DE SEGURIDAD" for new ingreso records
                    cargo = config.CARGO_DEFAULT if is_ingreso else lookup_data.get('CARGO')

                    # Birth date
                    fecha_nac = None
                    if fecha_nac_col and pd.notna(row.get(fecha_nac_col)):
                        fecha_nac = row.get(fecha_nac_col)
                    if pd.isna(fecha_nac):
                        fecha_nac = lookup_data.get('FECHA_NACIMIENTO_ASEGURADO')
                    fecha_nac = _format_date(fecha_nac)

                    # Movement date
                    fecha_novedad = row.get(fecha_nov_col) if fecha_nov_col else None
                    fecha_novedad = _format_date(fecha_novedad)

                    tipo_novedad = 'Ingreso' if is_ingreso else 'Retiro'

                    if is_ingreso:
                        all_ingresos_ids.add(n_id)
                    else:
                        all_retiros_ids.add(n_id)

                    processed_rows.append({
                        'NUMERO_POLIZA': config.NUMERO_POLIZA,
                        'APELLIDOS_ASEGURADO': apellidos,
                        'NOMBRES_ASEGURADO': nombres,
                        'TIPO_IDENTIFICACION_ASEGURADO': 'CC',
                        'NUMERO_IDENTIFICACION_ASEGURADO': n_id,
                        'CARGO': cargo,
                        'GENERO': genero,
                        'FECHA_NACIMIENTO_ASEGURADO': fecha_nac,
                        'VALOR_ASEGURADO_MUERTE': valor_aseg,
                        'TIPO_NOVEDAD': tipo_novedad,
                        'FECHA_NOVEDAD': fecha_novedad
                    })
                    count_records += 1

                logs.append(f"  -> Processed {count_records} valid records")

            except Exception as e:
                logs.append(f"  [!] Error processing {sheet_name}: {e}")
                logger.error(f"Error processing {fname} - {sheet_name}: {e}")

    # 3. Generate output DataFrames — both files use same 11 columns
    expected_cols = [
        'NUMERO_POLIZA', 'APELLIDOS_ASEGURADO', 'NOMBRES_ASEGURADO',
        'TIPO_IDENTIFICACION_ASEGURADO', 'NUMERO_IDENTIFICACION_ASEGURADO',
        'CARGO', 'GENERO', 'FECHA_NACIMIENTO_ASEGURADO', 'VALOR_ASEGURADO_MUERTE',
        'TIPO_NOVEDAD', 'FECHA_NOVEDAD'
    ]

    # ── Split ALL movements by cutoff date (26 of month before cobro month) ──
    # Anything strictly AFTER the cutoff goes to the "next month" file and does
    # NOT affect this month's movements or cobro plantilla.
    cutoff = _cobro_cutoff_date(cobro_mes, cobro_anio)

    current_rows: List[Dict[str, Any]] = []
    next_rows: List[Dict[str, Any]] = []
    for row in processed_rows:
        mov_date = _parse_full_date(row.get('FECHA_NOVEDAD'))
        if cutoff and mov_date and mov_date > cutoff:
            next_rows.append(row)
        else:
            current_rows.append(row)

    if cutoff:
        logs.append(f"Corte de cobro: {cutoff.strftime('%d/%m/%Y')} — "
                    f"{len(current_rows)} movimientos del mes, {len(next_rows)} para el mes siguiente")

    # Only movements within the current cutoff affect this month's master
    current_retiro_ids = {r['NUMERO_IDENTIFICACION_ASEGURADO']
                          for r in current_rows if r['TIPO_NOVEDAD'] == 'Retiro'}
    current_ingreso_ids = {r['NUMERO_IDENTIFICACION_ASEGURADO']
                           for r in current_rows if r['TIPO_NOVEDAD'] == 'Ingreso'}

    # Sort current movements: Ingresos first, then Retiros
    current_rows.sort(key=lambda r: (0 if r['TIPO_NOVEDAD'] == 'Ingreso' else 1))
    result_df = pd.DataFrame(current_rows, columns=expected_cols)
    next_month_df = pd.DataFrame(next_rows, columns=expected_cols)

    # Build master: keep people NOT retired this month, add this month's ingresos as Cobro
    new_master_rows = []
    seen_ids = set()
    for idx, row in master_df.iterrows():
        c_id = clean_id(row.get('NUMERO_IDENTIFICACION_ASEGURADO'))
        if not c_id or c_id == 'NAN' or c_id in current_retiro_ids or c_id in seen_ids:
            continue
        # Skip people already marked as Retiro in the master (no longer insured)
        if str(row.get('TIPO_NOVEDAD', '')).strip().lower() == 'retiro':
            continue
        row_dict = row.to_dict()
        row_dict['FECHA_NACIMIENTO_ASEGURADO'] = _format_date(row_dict.get('FECHA_NACIMIENTO_ASEGURADO'))
        # Everyone carried into the cobro plantilla is billed → TIPO_NOVEDAD = Cobro
        row_dict['TIPO_NOVEDAD'] = 'Cobro'
        new_master_rows.append(row_dict)
        seen_ids.add(c_id)

    # Add this month's ingresos to master with COBRO tipo_novedad for billing
    for row in current_rows:
        if row['TIPO_NOVEDAD'] == 'Ingreso':
            c_id = str(row.get('NUMERO_IDENTIFICACION_ASEGURADO', '')).strip()
            if c_id in seen_ids:
                continue
            master_row = dict(row)
            master_row['TIPO_NOVEDAD'] = 'Cobro'
            new_master_rows.append(master_row)
            seen_ids.add(c_id)

    new_master_df = pd.DataFrame(new_master_rows)
    for col in expected_cols:
        if col not in new_master_df.columns:
            new_master_df[col] = None
    new_master_df = new_master_df[expected_cols]

    logs.append("Processing completed successfully")

    # Create preview data - all records, no limit
    preview_movs = result_df.replace({np.nan: None}).to_dict(orient='records')
    preview_master = new_master_df.replace({np.nan: None}).to_dict(orient='records')

    return {
        'status': 'success',
        'statistics': {
            'total_movements': len(result_df),
            'total_ingresos': len(current_ingreso_ids),
            'total_retiros': len(current_retiro_ids),
            'total_next_month': len(next_month_df),
            'master_total_before': len(master_df),
            'master_total_after': len(new_master_df)
        },
        'dataframes': {
            'movements': result_df,
            'master': new_master_df,
            'next_month': next_month_df
        },
        'preview': {
            'movements': preview_movs,
            'master': preview_master
        },
        'logs': logs
    }


def generate_output_files(
    movements_df: pd.DataFrame,
    master_df: pd.DataFrame,
    output_dir: str,
    cobro_mes: int = 0,
    cobro_anio: int = 0,
    next_month_df: Optional[pd.DataFrame] = None
) -> Dict[str, Dict[str, str]]:
    """
    Write processed DataFrames to Excel files.

    The DataFrames arrive already split by cutoff date in process_excel_files:
    - movements_df: this month's movements (on/before the 26 cutoff)
    - master_df: this month's cobro plantilla
    - next_month_df: movements after the 26 cutoff (next billing cycle)

    Creates up to three Excel files:
    - INGRESOS_Y_RETIROS_GENERADOS_[MONTH]_[YEAR].xlsx
    - PLANTILLA_CARGUE_COBRO_[MONTH]_[YEAR].xlsx
    - MOVIMIENTOS_PARA_MES_SIGUIENTE_[MONTH]_[YEAR].xlsx (only if there are any)

    Returns:
        Dict[str, Dict[str, str]]: File metadata keyed by 'movements', 'master',
            and optionally 'next_month'; each has 'filename' and 'path'.
    """
    current_month: str = datetime.now().strftime('%B_%Y').upper()

    filenames = {
        'movements': f"INGRESOS_Y_RETIROS_GENERADOS_{current_month}.xlsx",
        'master': f"PLANTILLA_CARGUE_COBRO_{current_month}.xlsx",
        'next_month': f"MOVIMIENTOS_PARA_MES_SIGUIENTE_{current_month}.xlsx"
    }

    cols = [
        'NUMERO_POLIZA', 'APELLIDOS_ASEGURADO', 'NOMBRES_ASEGURADO',
        'TIPO_IDENTIFICACION_ASEGURADO', 'NUMERO_IDENTIFICACION_ASEGURADO',
        'CARGO', 'GENERO', 'FECHA_NACIMIENTO_ASEGURADO', 'VALOR_ASEGURADO_MUERTE',
        'TIPO_NOVEDAD', 'FECHA_NOVEDAD'
    ]

    # Calculate fecha_cobro: 26 of the month BEFORE cobro_mes
    fecha_cobro = None
    if cobro_mes and cobro_anio:
        prev_mes  = 12 if cobro_mes == 1 else cobro_mes - 1
        prev_anio = cobro_anio - 1 if cobro_mes == 1 else cobro_anio
        fecha_cobro = f"26/{prev_mes:02d}/{prev_anio}"

    # Apply fecha_cobro to cobro rows in the plantilla
    master_df_out = master_df.copy()
    if fecha_cobro:
        mask = master_df_out['TIPO_NOVEDAD'].astype(str).str.strip().str.lower() == 'cobro'
        master_df_out.loc[mask, 'FECHA_NOVEDAD'] = fecha_cobro

    try:
        movements_path = os.path.join(output_dir, filenames['movements'])
        master_path = os.path.join(output_dir, filenames['master'])

        _write_with_template(movements_df, cols, movements_path)
        _write_with_template(master_df_out, cols, master_path)

        result = {
            'movements': {'filename': filenames['movements'], 'path': movements_path},
            'master': {'filename': filenames['master'], 'path': master_path}
        }

        # Third file: movements after the 26th, reported next billing cycle
        if next_month_df is not None and len(next_month_df) > 0:
            next_month_path = os.path.join(output_dir, filenames['next_month'])
            _write_with_template(next_month_df, cols, next_month_path)
            result['next_month'] = {'filename': filenames['next_month'], 'path': next_month_path}
            logger.info(f"{len(next_month_df)} movimientos posteriores al 26 -> {filenames['next_month']}")

        logger.info(f"Output files generated: {filenames['movements']}, {filenames['master']}")
        return result
    except Exception as e:
        logger.error(f"Error generating output files: {e}")
        raise


def _cobro_cutoff_date(cobro_mes: int, cobro_anio: int) -> Optional[datetime]:
    """Return the cutoff datetime = 26 of the month BEFORE the cobro month."""
    if not cobro_mes or not cobro_anio:
        return None
    prev_mes  = 12 if cobro_mes == 1 else cobro_mes - 1
    prev_anio = cobro_anio - 1 if cobro_mes == 1 else cobro_anio
    return datetime(prev_anio, prev_mes, 26)


def _parse_full_date(date_val: Any) -> Optional[datetime]:
    """Parse a date value (DD/MM/YYYY, YYYY-MM-DD, Timestamp/datetime) into a datetime."""
    if date_val is None:
        return None
    if isinstance(date_val, (pd.Timestamp, datetime)):
        try:
            return datetime(date_val.year, date_val.month, date_val.day)
        except (ValueError, AttributeError):
            return None
    s = str(date_val).strip()
    if not s or s.lower() in ('nan', 'nat', 'none'):
        return None
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.strptime(s[:10], fmt)
        except ValueError:
            continue
    return None


def _write_with_template(df: pd.DataFrame, columns: List[str], output_path: str) -> None:
    """
    Write DataFrame to Sheet1 of a copy of PlantillaCargue template.
    Preserves 'Descripcion Campos' sheet unchanged.
    """
    template = getattr(config, 'TEMPLATE_PATH', None)

    if template and os.path.exists(template):
        shutil.copy2(template, output_path)
        wb = load_workbook(output_path)
        # Remove Sheet1 and recreate it completely clean
        if 'Sheet1' in wb.sheetnames:
            del wb['Sheet1']
        ws = wb.create_sheet('Sheet1')
        # Remove all data validations from the workbook to avoid system errors
        for sheet in wb.worksheets:
            sheet.data_validations.dataValidation = []
    else:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'

    # Ensure ALL expected columns exist in df
    df = df.copy()
    for col in columns:
        if col not in df.columns:
            df[col] = None

    # Numeric and date column types
    numeric_cols = {'VALOR_ASEGURADO_MUERTE', 'NUMERO_POLIZA'}
    date_cols    = {'FECHA_NACIMIENTO_ASEGURADO', 'FECHA_NOVEDAD'}

    # Write header row
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Write data rows with proper types
    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            val = _clean_cell(row.get(col_name))
            if val is not None:
                if col_name in numeric_cols:
                    try:
                        val = float(str(val).replace(',', '').strip())
                        val = int(val) if val == int(val) else val
                    except (ValueError, TypeError):
                        pass
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(output_path)


def _clean_cell(val: Any) -> Any:
    """Return None for any empty/null-like value so Excel cells are truly blank."""
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(val, float) and (val != val):  # NaN check
        return None
    s = str(val).strip()
    if s.lower() in ('nan', 'none', 'nat', 'n/a', ''):
        return None
    return val


# Helper functions

def _read_excel_safe(filepath: str) -> pd.DataFrame:
    """
    Read Excel file safely, trying different sheet names.

    Attempts to read 'Sheet1' first, then falls back to the last sheet
    in the workbook if Sheet1 is not found.

    Args:
        filepath: Path to the Excel file

    Returns:
        pd.DataFrame: The read Excel data

    Raises:
        FileNotFoundError: If file does not exist
        pd.errors.ParserError: If file cannot be parsed
    """
    try:
        return pd.read_excel(filepath, sheet_name="Sheet1")
    except ValueError:
        xls: pd.ExcelFile = pd.ExcelFile(filepath)
        return pd.read_excel(filepath, sheet_name=xls.sheet_names[-1])


def _find_header_row(df_raw: pd.DataFrame, max_rows: int) -> int:
    """
    Find header row by searching for identification-related keywords.

    Scans up to max_rows looking for common header keywords like
    IDENTIFICACION, C.C., CEDULA, etc.

    Args:
        df_raw: Raw DataFrame read without headers
        max_rows: Maximum number of rows to search

    Returns:
        int: 0-based index of header row, or -1 if not found
    """
    keywords: List[str] = ['IDENTIFICACION', 'C.C.', 'CEDULA', 'N° DOC', 'DOC']

    for i in range(min(max_rows, len(df_raw))):
        row_str: str = " ".join([str(x).upper() for x in df_raw.iloc[i].values])
        if any(kw in row_str for kw in keywords):
            return i

    return -1


def _find_column(columns: List[str], keywords: List[str]) -> Optional[str]:
    """
    Find first column matching any of the keywords.

    Performs a case-sensitive substring match against column names.

    Args:
        columns: List of column names to search
        keywords: List of keywords to match against

    Returns:
        str: First matching column name, or None if no match found
    """
    for col in columns:
        for kw in keywords:
            if kw in col:
                return col
    return None


_MESES_NUM = {
    'ENERO': '01', 'FEBRERO': '02', 'MARZO': '03', 'ABRIL': '04',
    'MAYO': '05', 'JUNIO': '06', 'JULIO': '07', 'AGOSTO': '08',
    'SEPTIEMBRE': '09', 'OCTUBRE': '10', 'NOVIEMBRE': '11', 'DICIEMBRE': '12'
}

def _format_date(date_val: Any) -> str:
    """Format date value as DD/MM/YYYY."""
    try:
        if date_val is None or (not isinstance(date_val, (pd.Timestamp, datetime)) and pd.isna(date_val)):
            return ""
    except (TypeError, ValueError):
        pass

    if isinstance(date_val, (pd.Timestamp, datetime)):
        return date_val.strftime('%d/%m/%Y')

    s = str(date_val).strip()
    if not s or s.lower() in ('nan', 'nat', 'none'):
        return ""

    # Already DD/MM/YYYY
    if len(s) == 10 and s[2] == '/' and s[5] == '/':
        return s

    # Convert YYYY-MM-DD → DD/MM/YYYY
    if len(s) == 10 and s[4] == '-' and s[7] == '-':
        parts = s.split('-')
        return f"{parts[2]}/{parts[1]}/{parts[0]}"

    # Convert Spanish text "01 DE DICIEMBRE DE 2025" → "01/12/2025"
    parts = s.upper().split()
    if len(parts) >= 4 and parts[1] == 'DE' and parts[3] == 'DE':
        dia = parts[0].zfill(2)
        mes = _MESES_NUM.get(parts[2])
        anio = parts[4] if len(parts) >= 5 else parts[3]
        if mes:
            return f"{dia}/{mes}/{anio}"

    return s


def _normalize_genero(value: Any) -> Optional[str]:
    """Normalize gender value to single letter 'M' or 'F' as required by PlantillaCargue."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    v = str(value).strip().upper()
    if v in ('M', 'MASCULINO', 'MALE', 'HOMBRE', 'H'):
        return 'M'
    if v in ('F', 'FEMENINO', 'FEMALE', 'MUJER'):
        return 'F'
    return None
